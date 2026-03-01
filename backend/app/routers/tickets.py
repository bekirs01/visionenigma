import io
import csv
import json
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, Request, Header, BackgroundTasks, UploadFile, File
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from pydantic import BaseModel
from app.db import get_db
from app.models import Ticket, Category, AiAnalysis, TicketAttachment
from app.schemas import TicketCreate, TicketRead, TicketUpdate, TicketsResponse, AnalyzeResponse, SuggestReplyResponse, TicketAttachmentRead
from app.repositories.ticket_repo import TicketRepository
from app.services.mock_ai import MockAIService
from app.auth import require_admin, require_admin_dep
from app.services.ai_agent import AIAgent
from app.services.device_extract import extract_device_model
from app.services.telegram_service import maybe_send_telegram_alert
from app.services.attachment_storage import save_attachment
from app.services.attachment_extract import extract_text_from_attachment
from app.config import get_settings

router = APIRouter(prefix="/api", tags=["tickets"])


class RequestCategoriesResponse(BaseModel):
    items: List[str]


@router.get("/tickets/request-categories", response_model=RequestCategoriesResponse)
def get_request_categories(
    request: Request,
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    view: Optional[str] = Query("open", description="open | answered"),
    sort: Optional[str] = Query(None, description="created_at_desc | created_at_asc | created_desc | created_asc | priority"),
    db: Session = Depends(get_db),
    _admin: bool = Depends(require_admin_dep),
):
    # Use the same "open/answered" filtering semantics as list endpoint.
    view_val = view if view in ("open", "answered") else "open"
    items = TicketRepository.get_request_categories(
        db,
        search=search,
        status=status,
        category_id=category_id,
        view=view_val,
    )
    # NOTE: sort is accepted for signature parity with admin panel,
    # but doesn't affect distinct set.
    return RequestCategoriesResponse(items=items)


def process_ticket_ai_background(ticket_id: int):
    """Фоновая обработка тикета AI-агентом (ручное создание). ai_status=done/failed."""
    from app.db import SessionLocal

    db = SessionLocal()
    try:
        settings = get_settings()
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            return
        if not settings.openai_api_key:
            ticket.ai_status = "done"
            ticket.ai_error = None
            db.commit()
            return

        print(f"[AI Background] Начинаем обработку тикета #{ticket_id}")
        agent = AIAgent(db)
        result = agent.process_ticket(ticket, attachments_summary="", attachments_extracted_text=(ticket.attachments_text or ""))
        agent.update_ticket_with_result(ticket, result)
        ticket.ai_status = "done"
        ticket.ai_error = None
        db.commit()
        print(f"[AI Background] Обработка тикета #{ticket_id} завершена")
        try:
            maybe_send_telegram_alert(db, ticket)
        except Exception as tg_err:
            print(f"[AI Background] Telegram alert skip: {tg_err}")
    except Exception as e:
        err_msg = str(e)[:500]
        print(f"[AI Background] Ошибка тикета #{ticket_id}: {err_msg}")
        try:
            ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
            if ticket:
                ticket.ai_status = "failed"
                ticket.ai_error = err_msg
                db.commit()
        except Exception:
            db.rollback()
    finally:
        db.close()


@router.get("/tickets", response_model=TicketsResponse)
def list_tickets(
    request: Request,
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    request_category: Optional[str] = Query(None),
    sort: Optional[str] = Query(None, description="created_at_desc | created_at_asc | created_desc | created_asc | priority"),
    view: Optional[str] = Query(None, description="open = только не отвеченные (по умолчанию), answered = только архив"),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    client_token: Optional[str] = Query(None, alias="client_token"),
    x_client_token: Optional[str] = Header(None, alias="X-Client-Token"),
    db: Session = Depends(get_db),
):
    token = (client_token or x_client_token or "").strip()
    view_val = view if view in ("open", "answered") else None
    repo = TicketRepository()
    if token:
        total = repo.get_count(db, search=search, status=status, category_id=category_id, request_category=request_category, client_token=token, view=view_val)
        tickets = repo.get_list(db, search=search, status=status, category_id=category_id, request_category=request_category, client_token=token, view=view_val, sort=sort, limit=limit, offset=offset)
        return TicketsResponse(items=tickets, total=total)
    if require_admin(request):
        total = repo.get_count(db, search=search, status=status, category_id=category_id, request_category=request_category, client_token=None, view=view_val or "open")
        tickets = repo.get_list(db, search=search, status=status, category_id=category_id, request_category=request_category, client_token=None, view=view_val or "open", sort=sort, limit=limit, offset=offset)
        return TicketsResponse(items=tickets, total=total)
    return TicketsResponse(items=[], total=0)


@router.post("/tickets", response_model=TicketRead)
def create_ticket(data: TicketCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    device_from_text = extract_device_model((data.subject or "") + "\n" + (data.body or ""))
    ticket = Ticket(
        sender_email=data.sender_email,
        sender_name=data.sender_name,
        subject=data.subject,
        body=data.body,
        status=data.status,
        priority=data.priority,
        category_id=data.category_id,
        source=data.source,
        client_token=(data.client_token or "").strip() or None,
        sender_full_name=data.sender_full_name,
        sender_phone=data.sender_phone,
        object_name=data.object_name,
        device_info=(data.device_info or "").strip() or None,
        device_type=device_from_text,
    )
    db.add(ticket)
    db.commit()
    db.refresh(ticket)

    # AI анализ в фоновом режиме - клиент не ждёт
    print(f"[AI] Тикет #{ticket.id} создан. Запуск фоновой обработки...")
    background_tasks.add_task(process_ticket_ai_background, ticket.id)

    return ticket


def _parse_serial_numbers(val) -> str:
    """Преобразует serial_numbers в строку для экспорта."""
    if not val:
        return ""
    if isinstance(val, list):
        return ", ".join(val)
    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            if isinstance(parsed, list):
                return ", ".join(parsed)
        except json.JSONDecodeError:
            pass
        return val
    return str(val)


# Заголовки для экспорта (ЭРИС кейс)
EXPORT_HEADERS = [
    "ID", "Дата", "ФИО", "Организация", "Телефон", "Email",
    "Заводские номера", "Тип прибора", "Эмоциональный окрас",
    "Категория запроса", "Суть вопроса", "Тема", "Статус",
    "Приоритет", "Требуется оператор"
]


def _ticket_to_export_row(t) -> list:
    """Преобразует тикет в строку для экспорта."""
    return [
        t.id,
        t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
        t.sender_full_name or t.sender_name or "",
        t.object_name or "",
        t.sender_phone or "",
        t.sender_email or "",
        _parse_serial_numbers(t.serial_numbers),
        t.device_type or "",
        {"positive": "Позитивный", "neutral": "Нейтральный", "negative": "Негативный"}.get(t.sentiment or "", ""),
        t.request_category or "",
        t.issue_summary or "",
        t.subject or "",
        {"not_completed": "Не завершён", "completed": "Завершён"}.get(t.status or "", t.status or ""),
        {"low": "Низкий", "medium": "Средний", "high": "Высокий"}.get(t.priority or "", t.priority or ""),
        "Да" if t.operator_required else "Нет"
    ]


@router.get("/tickets/export.csv")
def export_tickets_csv(
    request: Request,
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    request_category: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _admin: bool = Depends(require_admin_dep),
):
    """Экспорт тикетов в CSV с полными полями ЭРИС."""
    repo = TicketRepository()
    tickets = repo.get_list(
        db,
        search=search,
        status=status,
        category_id=category_id,
        request_category=request_category,
        client_token=None,
        view=None,
        limit=5000,
        offset=0,
    )

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(EXPORT_HEADERS)
    for t in tickets:
        writer.writerow(_ticket_to_export_row(t))
    body = "\ufeff" + output.getvalue()
    filename = f"tickets-{datetime.now().strftime('%Y-%m-%d')}.csv"
    return Response(
        content=body.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tickets/export.xlsx")
def export_tickets_xlsx(
    request: Request,
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    request_category: Optional[str] = Query(None),
    view: Optional[str] = Query(None, description="open | answered"),
    db: Session = Depends(get_db),
    _admin: bool = Depends(require_admin_dep),
):
    """Экспорт тикетов в XLSX с полными полями ЭРИС. Фильтры как в списке (view=open по умолчанию)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl не установлен. Выполните: pip install openpyxl"
        )

    view_val = view if view in ("open", "answered") else "open"
    repo = TicketRepository()
    tickets = repo.get_list(
        db,
        search=search,
        status=status,
        category_id=category_id,
        request_category=request_category,
        client_token=None,
        view=view_val,
        limit=5000,
        offset=0,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Обращения ЭРИС"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    for col, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Данные
    for row_idx, t in enumerate(tickets, 2):
        row_data = _ticket_to_export_row(t)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Ширина колонок
    column_widths = [8, 16, 25, 30, 15, 25, 20, 15, 15, 20, 40, 40, 12, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Закрепляем заголовок
    ws.freeze_panes = "A2"

    # Сохраняем в буфер
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"tickets-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tickets/{ticket_id}", response_model=TicketRead)
def get_ticket(
    ticket_id: int,
    request: Request,
    client_token: Optional[str] = Query(None),
    x_client_token: Optional[str] = Header(None, alias="X-Client-Token"),
    db: Session = Depends(get_db),
):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not require_admin(request):
        token = (client_token or x_client_token or "").strip()
        if not token or ticket.client_token != token:
            raise HTTPException(status_code=403, detail="Нет доступа к этому обращению")
    return ticket


ALLOWED_MIME_TYPES = {
    "application/pdf",
    "image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/plain", "text/csv",
    "application/octet-stream",
}
ALLOWED_EXTENSIONS = {
    ".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif",
    ".doc", ".docx", ".xls", ".xlsx", ".txt", ".csv",
}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB per file
MAX_FILES_PER_UPLOAD = 5


def _rerun_ai_with_attachments(ticket_id: int):
    """Re-process AI after new attachments are uploaded (web form)."""
    from app.db import SessionLocal

    db = SessionLocal()
    try:
        ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
        if not ticket:
            return
        atts = db.query(TicketAttachment).filter(TicketAttachment.ticket_id == ticket_id).all()
        if not atts:
            return

        from app.services.attachment_storage import UPLOADS_DIR
        summary_parts = []
        extracted_parts = []
        for att in atts:
            summary_parts.append(f"- {att.filename} ({att.mime_type}, {(att.size_bytes or 0) / 1024:.1f} KB)")
            file_path = UPLOADS_DIR / att.storage_path
            if file_path.exists():
                data = file_path.read_bytes()
                ok, text = extract_text_from_attachment(att.filename, att.mime_type, data)
                if text and text.strip():
                    extracted_parts.append(f"[{att.filename}]\n{text}")

        attachments_summary = "\n".join(summary_parts)
        attachments_extracted_text = "\n\n".join(extracted_parts)
        ticket.attachments_text = attachments_extracted_text or None

        settings = get_settings()
        if settings.openai_api_key:
            agent = AIAgent(db)
            result = agent.process_ticket(
                ticket,
                attachments_summary=attachments_summary,
                attachments_extracted_text=attachments_extracted_text,
            )
            agent.update_ticket_with_result(ticket, result)
            ticket.ai_status = "done"
            ticket.ai_error = None
        db.commit()
    except Exception as e:
        print(f"[Attachment AI] Error re-processing ticket #{ticket_id}: {e}")
        try:
            db.rollback()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/tickets/{ticket_id}/attachments", response_model=List[TicketAttachmentRead])
def upload_ticket_attachments(
    ticket_id: int,
    request: Request,
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    client_token: Optional[str] = Query(None),
    x_client_token: Optional[str] = Header(None, alias="X-Client-Token"),
    db: Session = Depends(get_db),
):
    """Upload file attachments to a ticket. Access: admin or owner via client_token."""
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not require_admin(request):
        token = (client_token or x_client_token or "").strip()
        if not token or ticket.client_token != token:
            raise HTTPException(status_code=403, detail="Нет доступа к этому обращению")

    if len(files) > MAX_FILES_PER_UPLOAD:
        raise HTTPException(status_code=400, detail=f"Максимум {MAX_FILES_PER_UPLOAD} файлов за раз")

    results: List[TicketAttachmentRead] = []
    base_url = str(request.base_url).rstrip("/")

    for f in files:
        fname = f.filename or "attachment"
        mime = (f.content_type or "application/octet-stream").lower()
        ext = fname.lower()[fname.rfind("."):] if "." in fname else ""
        if mime not in ALLOWED_MIME_TYPES and ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"Тип файла '{fname}' ({mime}) не поддерживается. Допустимые: PDF, JPG, PNG, DOCX, XLS, TXT, CSV.",
            )
        data = f.file.read()
        if len(data) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"Файл '{fname}' превышает лимит {MAX_FILE_SIZE // (1024 * 1024)} MB",
            )

        storage_path = save_attachment(ticket_id, fname, data)
        att = TicketAttachment(
            ticket_id=ticket_id,
            filename=fname,
            mime_type=mime,
            size_bytes=len(data),
            storage_path=storage_path,
        )
        db.add(att)
        db.flush()

        results.append(
            TicketAttachmentRead(
                id=att.id,
                ticket_id=att.ticket_id,
                filename=att.filename,
                mime_type=att.mime_type,
                size_bytes=att.size_bytes,
                storage_path=att.storage_path,
                created_at=att.created_at,
                download_url=f"{base_url}/uploads/{att.storage_path}",
            )
        )

    db.commit()

    background_tasks.add_task(_rerun_ai_with_attachments, ticket_id)

    return results


@router.get("/tickets/{ticket_id}/attachments", response_model=List[TicketAttachmentRead])
def get_ticket_attachments(
    ticket_id: int,
    request: Request,
    client_token: Optional[str] = Query(None),
    x_client_token: Optional[str] = Header(None, alias="X-Client-Token"),
    db: Session = Depends(get_db),
):
    """Список вложений тикета. Доступ: админ или владелец по client_token."""
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not require_admin(request):
        token = (client_token or x_client_token or "").strip()
        if not token or ticket.client_token != token:
            raise HTTPException(status_code=403, detail="Нет доступа к этому обращению")
    rows = db.query(TicketAttachment).filter(TicketAttachment.ticket_id == ticket_id).order_by(TicketAttachment.id).all()
    base_url = str(request.base_url).rstrip("/")
    return [
        TicketAttachmentRead(
            id=r.id,
            ticket_id=r.ticket_id,
            filename=r.filename,
            mime_type=r.mime_type,
            size_bytes=r.size_bytes,
            storage_path=r.storage_path,
            created_at=r.created_at,
            download_url=f"{base_url}/uploads/{r.storage_path}",
        )
        for r in rows
    ]


@router.delete("/tickets/{ticket_id}")
def delete_ticket(
    ticket_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _admin: bool = Depends(require_admin_dep),
):
    """Только админ может удалять тикеты. Пользователь не может удалять."""
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    db.delete(ticket)
    db.commit()
    return {"ok": True}


@router.patch("/tickets/{ticket_id}", response_model=TicketRead)
def update_ticket(
    ticket_id: int,
    data: TicketUpdate,
    request: Request,
    client_token: Optional[str] = Query(None),
    x_client_token: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not require_admin(request):
        token = (client_token or x_client_token or "").strip()
        if not token or ticket.client_token != token:
            raise HTTPException(status_code=403, detail="Нет доступа к этому обращению")
    if data.status is not None:
        ticket.status = data.status
    if data.priority is not None:
        ticket.priority = data.priority
    if data.category_id is not None:
        ticket.category_id = data.category_id
    if data.subject is not None:
        ticket.subject = data.subject
    if data.body is not None:
        ticket.body = data.body
    db.commit()
    db.refresh(ticket)
    return ticket


@router.post("/tickets/{ticket_id}/analyze", response_model=AnalyzeResponse)
def analyze_ticket(ticket_id: int, db: Session = Depends(get_db)):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    predicted_category, confidence = MockAIService.analyze(ticket.subject, ticket.body)
    analysis = AiAnalysis(
        ticket_id=ticket.id,
        predicted_category=predicted_category,
        confidence=confidence,
        provider="mock",
        model_version="mock-keyword-v1",
    )
    db.add(analysis)
    db.commit()
    db.refresh(analysis)
    return AnalyzeResponse(
        predicted_category=predicted_category,
        confidence=confidence,
        provider="mock",
        model_version="mock-keyword-v1",
        analysis_id=analysis.id,
    )


@router.post("/tickets/{ticket_id}/suggest-reply", response_model=SuggestReplyResponse)
def suggest_reply(ticket_id: int, db: Session = Depends(get_db)):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    category = ticket.category.name if ticket.category else "other"
    if not ticket.category:
        category, _ = MockAIService.analyze(ticket.subject, ticket.body)
    suggested = MockAIService.get_suggested_reply(category, ticket.subject, ticket.status)
    analysis = AiAnalysis(
        ticket_id=ticket.id,
        suggested_reply=suggested,
        provider="mock",
        model_version="mock-template-v1",
    )
    db.add(analysis)
    db.commit()
    db.refresh(analysis)
    return SuggestReplyResponse(
        suggested_reply=suggested,
        provider="mock",
        model_version="mock-template-v1",
        analysis_id=analysis.id,
    )
